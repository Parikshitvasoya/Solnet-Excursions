## importing the packages
from openpyxl import *
from tkinter import *
# importing subprocess
import subprocess

#loading workbook
wb = load_workbook('C:\\Users\\admin\\PycharmProjects\\Solent_Camp\\Data.xlsx')

sheet = wb.active

##defining function
def excel():
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 30
    sheet.column_dimensions['E'].width = 30
    sheet.column_dimensions['F'].width = 30
    sheet.column_dimensions['G'].width = 40

    sheet.cell(row=1, column=1).value = "Initial Location"
    sheet.cell(row=1, column=2).value = "Destination Location"
    sheet.cell(row=1, column=3).value = "Selected Vehicle"
    sheet.cell(row=1, column=4).value = "Total travelled person"
    sheet.cell(row=1, column=5).value = "Phone Number"
    sheet.cell(row=1, column=6).value = "Enter Email id"
    sheet.cell(row=1, column=7).value = "Enter your Location "


def focus1(event):
    course_field.focus_set()


def focus2(event):
    sem_field.focus_set()


def focus3(event):
    form_no_field.focus_set()


def focus4(event):
    contact_no_field.focus_set()


def focus5(event):
    email_id_field.focus_set()


def focus6(event):
    address_field.focus_set()


def clear():
    name_field.delete(0, END)
    course_field.delete(0, END)
    sem_field.delete(0, END)
    form_no_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_id_field.delete(0, END)
    address_field.delete(0, END)

## inserting command
def insert():
    if (name_field.get() == "" and
            course_field.get() == "" and
            sem_field.get() == "" and
            form_no_field.get() == "" and
            contact_no_field.get() == "" and
            email_id_field.get() == "" and
            address_field.get() == ""):

        print("empty input")
## add values in excel sheet
    else:
        current_row = sheet.max_row
        current_column = sheet.max_column

        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = course_field.get()
        sheet.cell(row=current_row + 1, column=3).value = sem_field.get()
        sheet.cell(row=current_row + 1, column=4).value = form_no_field.get()
        sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()
        sheet.cell(row=current_row + 1, column=6).value = email_id_field.get()
        sheet.cell(row=current_row + 1, column=7).value = address_field.get()

        # saving the file in csv format
        wb.save('C:\\Users\\admin\\PycharmProjects\\Solent_Camp\\Data.xlsx')

        name_field.focus_set()

        clear()

## delete function is called
def delete():
    subprocess.call(["python", "del.py"])

## mainfunction

if __name__ == "__main__":
    root = Tk()

    root.configure(background='lightgreen')

    root.title("Registration form")

    root.geometry("800x600")

    excel()

    heading = Label(root, text="Registration form", bg="yellow", height=4, width = 15)

    name = Label(root, text="Initial Location", bg="yellow", height=4, width = 15)

    course = Label(root, text="Destination Location", bg="yellow", height=4, width = 15)

    sem = Label(root, text="Selected Vehicle", bg="yellow", height=4, width = 15)

    form_no = Label(root, text="Total travelled person", bg="yellow", height=4, width = 15)

    contact_no = Label(root, text="Contact No.", bg="yellow", height=4, width = 15)

    email_id = Label(root, text="Enter your Email id", bg="yellow", height=4, width = 15)

    address = Label(root, text="Enter your Location", bg="yellow", height=4, width = 15)

    heading.grid(row=0, column=1)
    name.grid(row=1, column=0)
    course.grid(row=2, column=0)
    sem.grid(row=3, column=0)
    form_no.grid(row=4, column=0)
    contact_no.grid(row=5, column=0)
    email_id.grid(row=6, column=0)
    address.grid(row=7, column=0)

    name_field = Entry(root)
    course_field = Entry(root)
    sem_field = Entry(root)
    form_no_field = Entry(root)
    contact_no_field = Entry(root)
    email_id_field = Entry(root)
    address_field = Entry(root)

    name_field.bind("<Return>", focus1)

    course_field.bind("<Return>", focus2)

    sem_field.bind("<Return>", focus3)

    form_no_field.bind("<Return>", focus4)

    contact_no_field.bind("<Return>", focus5)

    email_id_field.bind("<Return>", focus6)

    name_field.grid(row=1, column=1, ipadx="100")
    course_field.grid(row=2, column=1, ipadx="100")
    sem_field.grid(row=3, column=1, ipadx="100")
    form_no_field.grid(row=4, column=1, ipadx="100")
    contact_no_field.grid(row=5, column=1, ipadx="100")
    email_id_field.grid(row=6, column=1, ipadx="100")
    address_field.grid(row=7, column=1, ipadx="100")

    excel()
    ## defining the buttons

    submit = Button(root, text="Confirm", fg="Blue", bg="yellow", command=insert, height=2, width = 5)
    submit.grid(row=8, column=1)

    excel()

    Delete = Button(root, text="Remove", fg="Blue", bg="yellow", command=delete, height=2, width = 5)
    Delete.grid(row=9, column=3)

    # processing the GUI
    root.mainloop()
