from tkinter import *
from openpyxl import *
from tkinter import messagebox

def delete1():
    wb = load_workbook('C:\\Users\\admin\\PycharmProjects\\Solent_Camp\\Data.xlsx')
    sheet = wb.active
    n=int(inputtxt.get())
    sheet.cell(row=n, column=1).value = ""
    sheet.cell(row=n, column=2).value = ""
    sheet.cell(row=n, column=3).value = ""
    sheet.cell(row=n, column=4).value = ""
    sheet.cell(row=n, column=5).value = ""
    sheet.cell(row=n, column=6).value = ""
    sheet.cell(row=n, column=7).value = ""
    print(sheet.cell(row=n, column=1).value)
    wb.save('C:\\Users\\admin\\PycharmProjects\\Solent_Camp\\Data.xlsx')
    messagebox.showinfo("showinfo", "Details deleted")

root = Tk()
root.configure(background='lightgreen')
root.title("Booking form")
root.geometry("600x400")

inputtxt = Entry(root)
inputtxt.pack()
b=Button(text="Remove", command=delete1)
b.pack()
root.mainloop()